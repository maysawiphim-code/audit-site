"""ตรวจไฟล์บันทึกของหัวหน้าทีม (ตารางรายชั่วโมงแยกวันที่ 1 / วันที่ 2)

โครงไฟล์ต่างกันไปตามไซต์ (คอลัมน์และแถวไม่ตรงกัน) จึงหาตำแหน่งจากหัวตารางเสมอ:
- แถวหัวตารางคือแถวที่มีคำว่า "ชม."  จำนวน "ชม." ในแถว = จำนวนวันที่จับ
- คอลัมน์ยอดรายชั่วโมงคือคอลัมน์ที่หัวเขียนว่า "รายชั่วโมง" โดยดูป้ายด้านบนว่าเป็น "คนผ่าน" หรือ "รถผ่าน"
- คอลัมน์ค่าดิปคือคอลัมน์ที่หัวเขียนว่า "%diff คน" / "%diff รถ" (ผลต่างระหว่างสองวันของชั่วโมงนั้น)

ค่าดิป = |วันที่ 2 − วันที่ 1| ÷ (วันที่ 1 + วันที่ 2)
"""
from __future__ import annotations

import io
import re

DIP = 0.20      # ค่าดิปเกินเท่านี้ = ต้องดู
TOL = 0.5       # ยอดรายชั่วโมงคลาดเคลื่อนได้ ±0.5


def _load(data: bytes, filename: str):
    if filename.lower().endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=data)
        sh = book.sheet_by_index(0)
        return [[sh.cell(r, c).value if sh.cell(r, c).ctype != 0 else None
                 for c in range(sh.ncols)] for r in range(sh.nrows)]
    import openpyxl.styles.fonts as _f
    _f.Font.family.max = 100          # บางไฟล์ใส่ค่า font family เกินมาตรฐาน
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(data), data_only=True).worksheets[0]
    return [[c.value for c in row] for row in ws.iter_rows()]


def _num(v):
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).replace(",", "").strip()
    return float(t) if re.fullmatch(r"-?\d*\.?\d+", t) else None


def _S(v):
    return "" if v is None else str(v).strip()


def parse_team_sheet(data: bytes, filename: str) -> dict:
    """คืนค่ารายชั่วโมง พร้อมยอดของทั้งสองวัน ค่าดิปในไฟล์ และค่าดิปที่คำนวณใหม่"""
    grid = _load(data, filename)
    hours, shift_names = [], []

    for r, row in enumerate(grid):
        hcols = [c for c, v in enumerate(row) if _S(v) == "ชม."]
        if not hcols:
            continue
        above = grid[r - 1] if r > 0 else []
        shift = next((_S(v) for v in above if _S(v) in ("เช้า", "บ่าย", "ดึก")),
                     f"ผลัดที่ {len(shift_names) + 1}")
        shift_names.append(shift)

        val_cols = []
        for c, v in enumerate(row):
            if _S(v) != "รายชั่วโมง":
                continue
            kind = ""
            for cc in range(c, -1, -1):
                t = _S(above[cc]) if cc < len(above) else ""
                if t in ("คนผ่าน", "รถผ่าน"):
                    kind = "คน" if t == "คนผ่าน" else "รถ"
                    break
            day = sum(1 for hc in hcols if hc < c) - 1
            if kind and day >= 0:
                val_cols.append((c, kind, day))

        dip_cols = {}
        for src in (row, above):
            for c, v in enumerate(src):
                m = re.match(r"%\s*diff\s*(คน|รถ)", _S(v))
                if m:
                    dip_cols.setdefault(m.group(1), c)

        n_days = len(hcols)
        rr = r + 1
        while rr < len(grid):
            h = _num(grid[rr][hcols[0]]) if hcols[0] < len(grid[rr]) else None
            if h is None or h != int(h) or not 1 <= h <= 24:
                break
            rec = {"shift": shift, "hour": int(h), "row": rr + 1}
            for kind in ("คน", "รถ"):
                vals = []
                for d in range(n_days):
                    col = next((c for c, k, dd in val_cols if k == kind and dd == d), None)
                    vals.append((_num(grid[rr][col]) if col is not None and col < len(grid[rr]) else None) or 0.0)
                rec[kind] = vals
                total = sum(vals)
                rec[f"ดิป_{kind}"] = (abs(vals[1] - vals[0]) / total) if len(vals) > 1 and total else None
                col = dip_cols.get(kind)
                rec[f"ดิปในไฟล์_{kind}"] = (_num(grid[rr][col])
                                            if col is not None and col < len(grid[rr]) else None)
            hours.append(rec)
            rr += 1

    return {"hours": hours, "shifts": shift_names,
            "n_days": max((len(h["คน"]) for h in hours), default=0)}


def report_hourly(rep) -> dict:
    """ตารางรายชั่วโมง (ค่าเฉลี่ย/วัน) จากชีตกราฟของไฟล์รายงาน"""
    for sheet in rep.sheets:
        grid = sheet.grid
        hr = next((r for r, row in enumerate(grid) if any(_S(v) == "ช่วงเวลา" for v in row)), None)
        if hr is None:
            continue
        labcol = next(c for c, v in enumerate(grid[hr]) if _S(v) == "ช่วงเวลา")
        pcol = next((c for c, v in enumerate(grid[hr]) if "เฉลี่ย" in _S(v)), labcol + 1)
        ccol = next((c for c, v in enumerate(grid[hr + 1]) if "จำนวนรถ" in _S(v)), None)
        out = {"sheet": sheet.name, "hours": [], "คน": [], "รถ": []}
        for r in range(hr + 1, len(grid)):
            lab = _S(grid[r][labcol]) if labcol < len(grid[r]) else ""
            if not re.match(r"^\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}", lab):
                continue
            out["hours"].append(lab)
            out["คน"].append(_num(grid[r][pcol]) or 0.0)
            out["รถ"].append((_num(grid[r][ccol]) if ccol is not None and ccol < len(grid[r]) else None) or 0.0)
        if out["hours"]:
            return out
    return {}


def analyse(team: dict, report: dict | None = None, threshold: float = DIP,
            gap_pct: float = 0.05) -> dict:
    """หาชั่วโมงที่ค่าดิปเกินเกณฑ์ ตรวจสูตรดิปในไฟล์ และเทียบยอดกับไฟล์รายงาน"""
    hours = team.get("hours") or []
    if not hours:
        return {}
    rep_hours = (report or {}).get("hours") or []
    rows, dips, formula_bad, gap = [], [], [], []

    for i, h in enumerate(hours):
        for kind in ("คน", "รถ"):
            d = h[f"ดิป_{kind}"]
            in_file = h[f"ดิปในไฟล์_{kind}"]
            vals = h[kind]
            avg = sum(vals) / len(vals) if vals else 0.0
            rep_v = report[kind][i] if report and i < len(report.get(kind, [])) else None
            row = {
                "ผลัด": h["shift"], "ชม.": h["hour"],
                "ช่วงเวลา": rep_hours[i] if i < len(rep_hours) else "",
                "ประเภท": kind,
                "วันที่ 1": vals[0] if vals else None,
                "วันที่ 2": vals[1] if len(vals) > 1 else None,
                "ค่าดิป %": None if d is None else round(d * 100, 2),
                "ดิปในไฟล์ %": None if in_file is None else round(in_file * 100, 2),
                "เฉลี่ย/วัน": round(avg, 2),
                "ไฟล์รายงาน": rep_v,
                "ต่างจากรายงาน": None if rep_v is None else round(avg - rep_v, 2),
                "เกินเกณฑ์": "🚩" if (d is not None and d > threshold) else "",
            }
            rows.append(row)
            if d is not None and d > threshold:
                dips.append(row)
            if d is not None and in_file is not None and abs(d - in_file) > 0.005:
                formula_bad.append(row)
            if rep_v is not None and abs(avg - rep_v) > max(TOL, abs(rep_v) * gap_pct):
                gap.append(row)

    return {"rows": rows, "dips": dips, "formula_bad": formula_bad, "gap": gap,
            "n_hours": len(hours), "has_report": bool(report)}
