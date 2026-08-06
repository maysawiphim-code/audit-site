"""ตรวจไฟล์รายงานจับตัวเลข (site count report) — แกนตรวจสอบ

กติกา
- ตรวจเฉพาะชีตที่มองเห็น (ข้ามชีตที่ซ่อน) และข้ามชีต "สรุป (2)"
- ตรวจเฉพาะเซลล์ที่ "ใส่สี" ไว้ในไฟล์ (พื้นขาว/ไม่มีสี = ไม่ตรวจ)
- จำนวนนับคลาดเคลื่อนได้ ±0.5
- หัวใจของการตรวจคือยอดรวมของคนและรถ ซึ่งแต่ละไฟล์มีจำนวนจุด/จำนวนวันไม่เท่ากัน
"""
from __future__ import annotations

import io
import re
import zipfile
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

TOL = 0.5        # คลาดเคลื่อนของจำนวนนับ
PTOL = 0.0015    # คลาดเคลื่อนของ %สัดส่วน
VALKEYS = ["ผลัดเช้า", "ผลัดบ่าย", "ผลัดดึก", "รวมทั้งวัน", "%สัดส่วน"]
SKIP_SHEETS = re.compile(r"^\s*สรุป\s*\(?\s*2\s*\)?\s*$")


# ---------------------------------------------------------------- utilities
def col_name(c: int) -> str:
    s, c = "", c + 1
    while c > 0:
        c, m = divmod(c - 1, 26)
        s = chr(65 + m) + s
    return s


def a1(r: int, c: int) -> str:
    return f"{col_name(c)}{r + 1}"


def S(v) -> str:
    return "" if v is None else str(v).strip()


def num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = S(v).replace(",", "")
    return float(t) if re.fullmatch(r"-?\d*\.?\d+", t) else None


def fmt(n) -> str:
    if n is None:
        return "-"
    return f"{round(n, 4):,g}"


def near(a, b, tol=TOL) -> bool:
    return abs(a - b) <= tol


# ---------------------------------------------------------------- data model
@dataclass
class Issue:
    sheet: str
    cell: str
    sev: str          # bad | warn
    title: str
    detail: str
    exempt_color: bool = False   # True = ตรวจแม้ช่องนั้นไม่ได้ใส่สี (ข้อความ/หัวเรื่อง/ยอดรวมชีตสรุป)


@dataclass
class Sheet:
    name: str
    hidden: bool
    grid: list
    fills: dict = field(default_factory=dict)   # "A1" -> "FFCC99"
    errors: dict = field(default_factory=dict)  # "A1" -> "#DIV/0!"


@dataclass
class Report:
    filename: str
    sheets: list
    issues: list = field(default_factory=list)
    passed: list = field(default_factory=list)
    skipped: list = field(default_factory=list)
    data_name: str = ""
    site: str = ""
    year: str = ""
    days: list = field(default_factory=list)
    totals: list = field(default_factory=list)      # ยอดรวมคน/รถ ที่ดึงได้
    key_totals: list = field(default_factory=list)  # ยอดรวมสำคัญไว้เทียบกับรูป
    categories: dict = field(default_factory=dict)  # หมวด -> รายการย่อยในชีต data
    numbers: list = field(default_factory=list)  # ตัวเลขทั้งหมดในตาราง data

    @property
    def bads(self):
        return [i for i in self.issues if i.sev == "bad"]

    @property
    def warns(self):
        return [i for i in self.issues if i.sev == "warn"]


# ---------------------------------------------------------------- readers
def read_workbook(data: bytes, filename: str) -> list:
    """อ่านไฟล์ .xls/.xlsx เป็น list[Sheet] พร้อมค่าสีพื้นและเซลล์สูตรผิดพลาด"""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _read_xlsx(data)
    return _read_xls(data)


def _read_xls(data: bytes) -> list:
    import xlrd

    book = xlrd.open_workbook(file_contents=data, formatting_info=True)
    cmap = book.colour_map
    out = []
    for sh in book.sheets():
        grid, fills, errors = [], {}, {}
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                v = cell.value
                if cell.ctype == 0:
                    v = None
                elif cell.ctype == 5:
                    errors[a1(r, c)] = xlrd.error_text_from_code.get(int(v), "#ERROR")
                    v = None
                row.append(v)
                xf = book.xf_list[sh.cell_xf_index(r, c)]
                bg = xf.background
                if bg.fill_pattern == 1:
                    rgb = cmap.get(bg.pattern_colour_index)
                    if rgb and rgb != (255, 255, 255):
                        fills[a1(r, c)] = "%02X%02X%02X" % rgb
            grid.append(row)
        out.append(Sheet(sh.name, sh.visibility != 0, grid, fills, errors))
    return out


def _read_xlsx(data: bytes) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    out = []
    for ws in wb.worksheets:
        grid, fills, errors = [], {}, {}
        for r, row in enumerate(ws.iter_rows()):
            vals = []
            for c, cell in enumerate(row):
                v = cell.value
                if isinstance(v, str) and v.startswith("#") and v.endswith(("!", "?")):
                    errors[a1(r, c)] = v
                    v = None
                vals.append(v)
                f = cell.fill
                if f is not None and f.patternType and f.patternType != "none":
                    rgb = getattr(f.start_color, "rgb", None)
                    if isinstance(rgb, str) and len(rgb) >= 6:
                        rgb = rgb[-6:].upper()
                        if rgb != "FFFFFF":
                            fills[a1(r, c)] = rgb
            grid.append(vals)
        out.append(Sheet(ws.title, ws.sheet_state != "visible", grid, fills, errors))
    return out


# ---------------------------------------------------------------- parse "data"
def parse_data(grid: list) -> dict:
    """แยกโครงตารางสรุปผลในชีต data: หัวรายงาน / กลุ่มคอลัมน์ / แถวข้อมูล"""
    out = {"meta": [], "groups": [], "tables": [], "hrow": -1, "subrow": -1}

    for r in range(min(len(grid), 20)):
        lab = S(grid[r][0] if grid[r] else "")
        if not lab.endswith(":"):
            continue
        val, vc = "", 1
        for c in range(1, len(grid[r])):
            if S(grid[r][c]):
                val, vc = S(grid[r][c]), c
                break
        out["meta"].append({"row": r, "col": vc, "label": lab[:-1].strip(), "value": val})
        for c in range(1, len(grid[r])):          # คู่ที่สองในบรรทัดเดียวกัน
            t = S(grid[r][c])
            if len(t) > 1 and t.endswith(":"):
                v2, v2c = "", c + 1
                for d in range(c + 1, len(grid[r])):
                    if S(grid[r][d]):
                        v2, v2c = S(grid[r][d]), d
                        break
                out["meta"].append({"row": r, "col": v2c, "label": t[:-1].strip(), "value": v2})
                break

    for r, row in enumerate(grid):
        if any(S(v) == "หัวข้อ" for v in row):
            out["hrow"] = r
            break
    if out["hrow"] < 0:
        return out
    for r in range(out["hrow"], min(out["hrow"] + 3, len(grid))):
        if any(S(v) in VALKEYS for v in grid[r]):
            out["subrow"] = r
            break
    if out["subrow"] < 0:
        return out

    sub = grid[out["subrow"]]
    valcols = [(c, S(v)) for c, v in enumerate(sub) if S(v) in VALKEYS]
    if not valcols:
        return out

    groups, cur = [], None
    for c, key in valcols:
        if cur and c == cur["end"] + 1 and key not in [k for _, k in cur["keys"]]:
            cur["end"] = c
            cur["keys"].append((c, key))
        else:
            cur = {"start": c, "end": c, "keys": [(c, key)]}
            groups.append(cur)

    prev_end = -1
    for gr in groups:
        title = ""
        for c in range(gr["start"], max(prev_end, 0) - 1, -1):
            t = S(grid[out["hrow"]][c]) if c < len(grid[out["hrow"]]) else ""
            if t and t != "หัวข้อ":
                title = t
                break
        gr["title"] = title or "ผลรวม"
        ls, le = prev_end + 1, gr["start"] - 1
        gr["labelcols"] = (ls, le) if le >= ls else None
        prev_end = gr["end"]
    last = None
    for gr in groups:
        if gr["labelcols"]:
            last = gr["labelcols"]
        else:
            gr["labelcols"] = last
    out["groups"] = groups

    tables = {}
    for gr in groups:
        tables.setdefault(gr["labelcols"], {"labelcols": gr["labelcols"], "groups": [], "rows": []})["groups"].append(gr)

    for t in tables.values():
        if not t["labelcols"]:
            continue
        ls, le = t["labelcols"]
        used = []
        for c in range(ls, le + 1):
            if any(c < len(grid[r]) and S(grid[r][c]) for r in range(out["subrow"] + 1, len(grid))):
                used.append(c)
        if not used:
            continue
        block_col, carry = used[0], {}
        for r in range(out["subrow"] + 1, len(grid)):
            ne = [(c, S(grid[r][c])) for c in used if c < len(grid[r]) and S(grid[r][c])]
            has_val = any(num(grid[r][c]) is not None
                          for gr in t["groups"] for c, _ in gr["keys"] if c < len(grid[r]))
            for c, v in ne:
                carry[c] = v
                for c2 in used:
                    if c2 > c:
                        carry[c2] = ""
            if not ne:
                continue
            lc, item = ne[-1]
            if lc == block_col and not has_val:
                continue
            is_total = "ยอดรวม" in item
            sub_lab = ""
            if lc != block_col and not is_total:
                for c in [c for c in used if block_col < c < lc][::-1]:
                    if carry.get(c):
                        sub_lab = carry[c]
                        break
            t["rows"].append({"r": r, "block": carry.get(block_col, ""), "sub": sub_lab,
                              "item": item, "is_total": is_total, "labelcol": lc})
    out["tables"] = list(tables.values())
    return out


# ---------------------------------------------------------------- checks
def check_data(sheet: Sheet, rep: Report):
    grid, name = sheet.grid, sheet.name
    D = parse_data(grid)
    add = lambda sev, r, c, title, detail: rep.issues.append(Issue(name, a1(r, c), sev, title, detail))
    seen = set()

    def ok_once(t):
        if t not in seen:
            seen.add(t)
            rep.passed.append((name, t))

    for m in D["meta"]:
        if m["value"] == "":
            add("bad", m["row"], m["col"], "ข้อมูลหัวรายงานว่าง", f'"{m["label"]}" ยังไม่ได้กรอก')
        if re.search(r"ชื่อ\s*SITE", m["label"], re.I):
            rep.site = re.sub(r"\(.*$", "", m["value"]).strip()
        if "จับตัวเลขวันที่" in m["label"]:
            y = re.search(r"25\d{2}", m["value"])
            if y:
                rep.year = y.group()
            d = re.search(r"ที่\s*(\d{1,2})", m["value"])
            if d:
                rep.days.append(d.group(1))

    if not D["tables"]:
        add("bad", 0, 0, "ไม่พบตารางสรุปผล", 'หาแถวหัวตาราง "หัวข้อ" ไม่เจอ จึงตรวจตัวเลขในชีตนี้ไม่ได้')
        return

    for t in D["tables"]:
        rows = t["rows"]
        for gr in t["groups"]:
            keycol = {k: c for c, k in gr["keys"]}
            shifts = [k for k in ("ผลัดเช้า", "ผลัดบ่าย", "ผลัดดึก") if k in keycol]

            # 1) ผลัดเช้า+บ่าย+ดึก = รวมทั้งวัน
            if shifts and "รวมทั้งวัน" in keycol:
                n = bad = 0
                for row in rows:
                    parts = [num(grid[row["r"]][keycol[k]]) for k in shifts]
                    tot = num(grid[row["r"]][keycol["รวมทั้งวัน"]])
                    if tot is None or any(p is None for p in parts):
                        continue
                    n += 1
                    s = sum(parts)
                    if not near(s, tot):
                        bad += 1
                        add("bad", row["r"], keycol["รวมทั้งวัน"], "ผลรวมผลัดไม่ตรง",
                            f'{gr["title"]} · {row["block"]} {row["item"]} : '
                            f'{" + ".join(fmt(p) for p in parts)} = {fmt(s)} แต่ช่องรวมทั้งวันใส่ {fmt(tot)}')
                if n and not bad:
                    ok_once(f'{gr["title"]}: ผลัดเช้า+บ่าย+ดึก = รวมทั้งวัน ครบ {n} แถว')

            # 2) ผลรวมกลุ่มย่อย = ยอดรวมของบล็อก + 3) %สัดส่วน
            totals = {r["block"]: r for r in rows if r["is_total"]}
            bysub = {}
            for r in rows:
                if not r["is_total"] and r["sub"]:
                    bysub.setdefault((r["block"], r["sub"]), []).append(r)
            for (blk, sub_lab), items in bysub.items():
                tot_row = totals.get(blk)
                if tot_row:
                    for key, c in keycol.items():
                        if key == "%สัดส่วน":
                            continue
                        tv = num(grid[tot_row["r"]][c])
                        if tv is None:
                            continue
                        parts = [num(grid[r["r"]][c]) for r in items]
                        parts = [p for p in parts if p is not None]
                        if len(parts) < 2:
                            continue
                        s = sum(parts)
                        if not near(s, tv):
                            add("bad", tot_row["r"], c, "ยอดรวมกลุ่มไม่ตรง",
                                f'{gr["title"]} · {blk} → {sub_lab} ({key}) : รวมรายการย่อยได้ {fmt(s)} '
                                f'แต่ช่องยอดรวมใส่ {fmt(tv)} (ต่าง {fmt(s - tv)})')
                        else:
                            ok_once(f'{blk} → {sub_lab} ({gr["title"]} · {key}) รวมได้ {fmt(s)} ตรงกับยอดรวม')
                if "%สัดส่วน" in keycol and "รวมทั้งวัน" in keycol:
                    base = sum(v for v in (num(grid[r["r"]][keycol["รวมทั้งวัน"]]) for r in items) if v is not None)
                    if base:
                        for r in items:
                            v = num(grid[r["r"]][keycol["รวมทั้งวัน"]])
                            p = num(grid[r["r"]][keycol["%สัดส่วน"]])
                            if v is None or p is None:
                                continue
                            if abs(p - v / base) > PTOL:
                                add("warn", r["r"], keycol["%สัดส่วน"], "%สัดส่วนไม่ตรง",
                                    f'{gr["title"]} · {blk} {r["item"]} : {fmt(v)}/{fmt(base)} '
                                    f'ควรเป็น {v / base * 100:.2f}% แต่ใส่ {p * 100:.2f}%')

            # 4) บล็อก "ทั้งหมด" = ผลรวมทุกบล็อกย่อย (รายรายการ)
            grand_blk = next((b for b in totals if "ทั้งหมด" in b), None)
            if grand_blk:
                src = {}
                for r in rows:
                    if not r["is_total"] and r["block"] != grand_blk and r["block"] in totals:
                        src.setdefault(r["item"], []).append(r)
                for r in [x for x in rows if not x["is_total"] and x["block"] == grand_blk]:
                    lst = src.get(r["item"], [])
                    if len(lst) < 2:
                        continue
                    for key, c in keycol.items():
                        if key == "%สัดส่วน":
                            continue
                        gv = num(grid[r["r"]][c])
                        if gv is None:
                            continue
                        vals = [v for v in (num(grid[x["r"]][c]) for x in lst) if v is not None]
                        s = sum(vals)
                        if not near(s, gv):
                            add("bad", r["r"], c, "ยอดรวมใหญ่ไม่ตรง",
                                f'{gr["title"]} · {grand_blk} {r["item"]} ({key}) : รวมจากทุกจุดได้ {fmt(s)} แต่ใส่ {fmt(gv)}')

            # 5) ★ ยอดรวมคน/รถทั้งหมด = ผลรวมยอดรวมย่อยที่ใช้ "สีเดียวกัน"
            tot_rows = [r for r in rows if r["is_total"]]
            if len(tot_rows) >= 2:
                kinds = {}
                for r in tot_rows:
                    kind = "คน" if "คน" in r["item"] else ("รถ" if "รถ" in r["item"] else "อื่น")
                    color = sheet.fills.get(a1(r["r"], r["labelcol"]), "")
                    kinds.setdefault(kind, []).append((r, color))
                for kind, lst in kinds.items():
                    if len(lst) < 2:
                        continue
                    bycolor = {}
                    for r, col in lst:
                        bycolor.setdefault(col, []).append(r)
                    grand_color = next((c for c, v in bycolor.items()
                                        if len(v) == 1 and any(len(v2) > 1 for c2, v2 in bycolor.items() if c2 != c)), None)
                    if not grand_color:
                        continue
                    grand_row = bycolor[grand_color][0]
                    parts = [r for r, col in lst if col != grand_color]
                    if len(parts) < 2:
                        continue
                    for key, c in keycol.items():
                        if key == "%สัดส่วน":
                            continue
                        gv = num(grid[grand_row["r"]][c])
                        if gv is None:
                            continue
                        vals = [v for v in (num(grid[r["r"]][c]) for r in parts) if v is not None]
                        if len(vals) < 2:
                            continue
                        s = sum(vals)
                        label = f'ยอดรวม{kind}ทั้งหมด ({gr["title"]} · {key})'
                        if not near(s, gv):
                            add("bad", grand_row["r"], c, f"ยอดรวม{kind}ทั้งหมดไม่ตรง",
                                f'{gr["title"]} ({key}) : รวมยอดของทุกจุด ({" + ".join(fmt(v) for v in vals)}) = {fmt(s)} '
                                f'แต่ช่องยอดรวมใส่ {fmt(gv)} — ต่าง {fmt(s - gv)}')
                        else:
                            ok_once(f"{label} ตรงกับผลรวมทุกจุด: {fmt(gv)}")
                            rep.totals.append({"kind": kind, "group": gr["title"], "key": key,
                                               "value": gv, "cell": a1(grand_row["r"], c)})

            # เก็บรายการย่อยของแต่ละหมวด (เพศ/อายุ/อาชีพ/ทิศ/ประเภทรถ)
            for r in rows:
                if r["sub"] and not r["is_total"] and r["item"]:
                    rep.categories.setdefault(r["sub"], set()).add(r["item"])

            # เก็บยอดรวมสำคัญไว้เทียบกับรูปในชีตแผนที่/กราฟ
            if "รวมทั้งวัน" in keycol:
                for r in rows:
                    if not r["is_total"]:
                        continue
                    v = num(grid[r["r"]][keycol["รวมทั้งวัน"]])
                    if v is None or v < 2:
                        continue
                    label = f'{r["block"]} {r["item"]}'.strip()
                    if not any(near(k["value"], v) and k["label"] == label for k in rep.key_totals):
                        rep.key_totals.append({"label": label, "value": v,
                                               "group": gr["title"], "cell": a1(r["r"], keycol["รวมทั้งวัน"])})

            # เก็บตัวเลข + ค่าติดลบ
            for row in rows:
                for key, c in keycol.items():
                    v = num(grid[row["r"]][c])
                    if v is None:
                        continue
                    rep.numbers.append(v)
                    if v < 0:
                        add("bad", row["r"], c, "ค่าติดลบ",
                            f'{gr["title"]} · {row["block"]} {row["item"]} ({key}) = {fmt(v)}')

        # 6) คอลัมน์เฉลี่ยของหลายวัน
        days = [g for g in t["groups"] if "วัน" in g["title"]]
        avg = next((g for g in t["groups"] if "เฉลี่ย" in g["title"]), None)
        if avg and len(days) >= 2:
            n = bad = 0
            for c, key in avg["keys"]:
                if key == "%สัดส่วน":
                    continue
                for row in rows:
                    av = num(grid[row["r"]][c])
                    if av is None:
                        continue
                    vals = []
                    for dg in days:
                        dc = next((cc for cc, kk in dg["keys"] if kk == key), None)
                        vals.append(num(grid[row["r"]][dc]) if dc is not None else None)
                    if any(v is None for v in vals):
                        continue
                    n += 1
                    m = sum(vals) / len(vals)
                    if not near(m, av):
                        bad += 1
                        add("bad", row["r"], c, "ค่าเฉลี่ยไม่ตรง",
                            f'{row["block"]} {row["item"]} ({key}) : เฉลี่ยจาก {", ".join(fmt(v) for v in vals)} '
                            f'ควรเป็น {fmt(m)} แต่ใส่ {fmt(av)}')
            if n and not bad:
                ok_once(f"คอลัมน์เฉลี่ยคำนวณถูกต้องครบ {n} ช่อง")


def check_hourly(sheet: Sheet, rep: Report) -> bool:
    grid, name = sheet.grid, sheet.name
    hr = next((r for r, row in enumerate(grid) if any("ช่วงเวลา" in S(v) for v in row)), None)
    if hr is None:
        return False
    labcol = next(c for c, v in enumerate(grid[hr]) if "ช่วงเวลา" in S(v))
    is_hour = lambda v: re.match(r"^\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}", S(v))
    hours, tots, shifts = [], [], []
    for r in range(hr + 1, len(grid)):
        lab = S(grid[r][labcol]) if labcol < len(grid[r]) else ""
        if is_hour(lab):
            hours.append(r)
        elif lab.startswith("รวม"):
            tots.append(r)
        elif "ผลัด" in lab:
            shifts.append(r)
    if not hours or not tots:
        return False
    add = lambda sev, r, c, title, detail: rep.issues.append(Issue(name, a1(r, c), sev, title, detail))
    width = max(len(grid[r]) for r in hours + tots)
    cols = bad = 0
    for c in range(labcol + 1, width):
        get = lambda rs: [v for v in (num(grid[r][c]) if c < len(grid[r]) else None for r in rs) if v is not None]
        hv, tv = get(hours), get(tots)
        if not hv and not tv:
            continue
        cols += 1
        head = S(grid[hr][c]) if c < len(grid[hr]) else col_name(c)
        if not near(sum(hv), sum(tv)):
            bad += 1
            add("bad", tots[-1], c, "ยอดรวมรายชั่วโมงไม่ตรง",
                f'คอลัมน์ "{head}" : รวมทุกช่วงเวลาได้ {fmt(sum(hv))} แต่แถวรวมใส่ {fmt(sum(tv))} '
                f'(ต่าง {fmt(sum(hv) - sum(tv))})')
        sv = get(shifts)
        if sv and not near(sum(sv), sum(tv)):
            add("warn", shifts[-1], c, "ยอดรวมผลัดไม่ตรง",
                f'คอลัมน์ "{head}" : รวมทุกผลัดได้ {fmt(sum(sv))} แต่แถวรวมใส่ {fmt(sum(tv))}')
        for r in tots:
            v = num(grid[r][c]) if c < len(grid[r]) else None
            if v is None or abs(v) < 2:
                continue
            if not any(near(b, v) for b in rep.numbers):
                add("warn", r, c, "ยอดรวมไม่พบในชีต data",
                    f'คอลัมน์ "{head}" ยอดรวม {fmt(v)} ไม่ปรากฏในตารางสรุปผลของชีต data')
    if cols and not bad:
        rep.passed.append((name, f"ตารางรายชั่วโมงรวมยอดตรงครบ {cols} คอลัมน์"))
    return True


def check_summary(sheet: Sheet, rep: Report):
    grid, name = sheet.grid, sheet.name
    add = lambda sev, r, c, title, detail: rep.issues.append(Issue(name, a1(r, c), sev, title, detail))
    checked = bad = 0
    for r, row in enumerate(grid):
        for c, raw in enumerate(row):
            if isinstance(raw, (int, float)) and not isinstance(raw, bool):
                v = float(raw)
                if abs(v) < 2:
                    continue
                checked += 1
                if not any(near(b, v) for b in rep.numbers):
                    bad += 1
                    add("bad", r, c, "ตัวเลขไม่ตรงกับชีต data",
                        f"ค่า {fmt(v)} ในชีตนี้ไม่มีอยู่ในตารางสรุปผลของชีต data — "
                        f"อาจเป็นตัวเลขค้างจากงานเดิมหรือคำนวณคนละฐาน")
                continue
            t = S(raw)
            if len(t) < 6:
                continue
            if re.search(r"…|\.{5,}", t):
                add("warn", r, c, "ยังไม่ได้เติมข้อความ", f'ข้อความสรุปยังเป็นจุดไข่ปลา: "{t[:60]}"')
            for ns in re.findall(r"\d[\d,]*\.?\d*", t):
                v = float(ns.replace(",", ""))
                if v < 100:
                    continue
                if f"{ns}%" in t:
                    if not any(near(b, v / 100, PTOL) for b in rep.numbers):
                        add("warn", r, c, "ตัวเลขในข้อความไม่ตรง",
                            f"ข้อความระบุ {ns}% แต่ไม่พบสัดส่วนนี้ในชีต data")
                    continue
                if 2500 < v < 2600:
                    continue
                if not any(near(b, v) for b in rep.numbers):
                    add("bad", r, c, "ตัวเลขในข้อความไม่ตรง",
                        f"ข้อความสรุประบุ {ns} แต่ไม่พบตัวเลขนี้ในชีต data")
    if checked and not bad:
        rep.passed.append((name, f"ตัวเลข {checked} ค่าตรงกับชีต data ทั้งหมด"))


def check_summary_tables(sheet: Sheet, rep: Report):
    """ตรวจตารางในชีตสรุป: ยอดรวมเพี้ยนเล็กน้อย · ยอดรวมในชีตเดียวกันขัดกันเอง · รายการย่อยหาย"""
    grid, name = sheet.grid, sheet.name
    add = lambda sev, r, c, title, detail: rep.issues.append(
        Issue(name, a1(r, c), sev, title, detail, exempt_color=True))

    # 1) ตัวเลขที่ "เกือบ" เท่ายอดรวมในชีต data (ต่างไม่เกิน 2%) = น่าจะคัดลอกยอดผิด
    for r, row in enumerate(grid):
        for c, v in enumerate(row):
            if not isinstance(v, (int, float)) or isinstance(v, bool) or abs(v) < 10:
                continue
            v = float(v)
            if any(near(b, v) for b in rep.numbers):   # มีค่านี้จริงในชีต data
                continue
            for k in rep.key_totals:
                if abs(v - k["value"]) <= max(abs(k["value"]) * 0.02, 1):
                    add("bad", r, c, "ยอดรวมในชีตสรุปไม่ตรงกับชีต data",
                        f'ค่า {fmt(v)} ในชีตนี้ใกล้เคียงกับ "{k["label"]}" ของชีต {rep.data_name} '
                        f'ซึ่งเป็น {fmt(k["value"])} (เซลล์ {k["cell"]}) — ต่าง {fmt(v - k["value"])}')
                    break

    # 2) ยอดรวมชนิดเดียวกันในชีตเดียวกันต้องเท่ากัน
    tot_cells = {}
    for r, row in enumerate(grid):
        label = " ".join(S(v) for v in row if isinstance(v, str))
        if not label:
            continue
        kind = "คน" if "คน" in label else ("รถ" if "รถ" in label else None)
        if kind and re.search(r"ยอดรวม|^เพศ|ทิศคน|ทิศรถ", label):
            for c, v in enumerate(row):
                if isinstance(v, (int, float)) and not isinstance(v, bool) and abs(v) >= 10:
                    tot_cells.setdefault(kind, []).append((r, c, float(v), label[:40]))
                    break
    for kind, lst in tot_cells.items():
        if len(lst) < 2:
            continue
        base = lst[0]
        for r, c, v, lab in lst[1:]:
            if not near(v, base[2]):
                add("bad", r, c, "ยอดรวมในชีตเดียวกันขัดกันเอง",
                    f'"{lab}" = {fmt(v)} แต่ "{base[3]}" ในชีตเดียวกัน = {fmt(base[2])} '
                    f"— ยอดรวมของ{kind}ต้องเป็นค่าเดียวกันทั้งชีต")

    # 3) รายการย่อยในตารางสรุปต้องครบเท่าชีต data
    import difflib
    texts = [S(v) for row in grid for v in row if isinstance(v, str) and S(v)]
    same = lambda a, b: difflib.SequenceMatcher(None, a, b).ratio() >= 0.8
    for cat, items in rep.categories.items():
        hit = {i for i in items if any(same(i, t) for t in texts)}
        missing = items - hit
        if len(hit) >= 2 and missing:
            add("bad", 0, 0, "ตารางสรุปมีรายการไม่ครบ",
                f'หมวด "{cat}" ในชีต {rep.data_name} มี {len(items)} รายการ '
                f'แต่ชีตนี้ขาด: {", ".join(sorted(missing))} — ยอดรวมของหมวดนี้จึงน้อยกว่าความจริง')


MONTHS = "มกราคม|กุมภาพันธ์|มีนาคม|เมษายน|พฤษภาคม|มิถุนายน|กรกฎาคม|สิงหาคม|กันยายน|ตุลาคม|พฤศจิกายน|ธันวาคม"


def check_titles(sheet: Sheet, rep: Report):
    grid, name = sheet.grid, sheet.name
    add = lambda sev, r, c, title, detail: rep.issues.append(
        Issue(name, a1(r, c), sev, title, detail, exempt_color=True))
    norm = lambda t: re.sub(r"[\s()（）]", "", S(t))
    for r in range(min(len(grid), 6)):
        for c, raw in enumerate(grid[r]):
            t = S(raw)
            if len(t) < 10:
                continue
            if rep.site and re.search(r"ทำเลเป้าหมาย|ตารางสรุปผล|แผนที่", t):
                key = norm(rep.site)[:10]
                if len(key) >= 4 and key not in norm(t):
                    add("bad", r, c, "ชื่อทำเลในหัวเรื่องไม่ตรง",
                        f'หัวเรื่องเขียนว่า "{t[:60]}" แต่ชีต data ระบุทำเลว่า "{rep.site}"')
            if rep.year and re.search(f"วัน|{MONTHS}", t):
                for y in re.findall(r"(?<![\d.])25\d{2}(?![\d.])", t):
                    if y != rep.year:
                        add("bad", r, c, "ปี พ.ศ. ไม่ตรง",
                            f"ข้อความระบุปี {y} แต่วันที่จับตัวเลขในชีต data เป็นปี {rep.year}")
            if rep.days and t.startswith("จับตัวเลขวัน"):
                d = [re.sub(r"\D", "", x) for x in re.findall(r"ที่\s*\d{1,2}", t)]
                if d and (len(d) != len(rep.days) or any(x not in rep.days for x in d)):
                    add("bad", r, c, "วันที่จับตัวเลขไม่ตรง",
                        f'ชีตนี้ระบุวันที่ {", ".join(d)} แต่ชีต data ระบุวันที่ {", ".join(rep.days)}')


# ---------------------------------------------------------------- main audit
def audit(data: bytes, filename: str) -> Report:
    sheets = read_workbook(data, filename)
    rep = Report(filename=filename, sheets=[])
    visible = []
    for sh in sheets:
        if sh.hidden:
            rep.skipped.append((sh.name, "ซ่อนไว้"))
        elif SKIP_SHEETS.match(sh.name):
            rep.skipped.append((sh.name, "ตั้งค่าไม่ต้องตรวจ"))
        else:
            visible.append(sh)
    rep.sheets = visible

    data_sheet = next((s for s in visible if s.name.lower() == "data"), None)
    if data_sheet is None:
        data_sheet = next((s for s in visible if any("หัวข้อ" == S(v) for row in s.grid for v in row)), None)
    if data_sheet is None:
        rep.issues.append(Issue("-", "", "bad", "ไม่พบชีต data", "ไฟล์นี้ไม่มีชีตข้อมูลหลักให้ใช้เป็นฐานเทียบ"))
    else:
        rep.data_name = data_sheet.name
        check_data(data_sheet, rep)

    for sh in visible:
        for cell, err in sh.errors.items():
            rep.issues.append(Issue(sh.name, cell, "bad", "สูตรคำนวณผิดพลาด",
                                    f"เซลล์แสดงค่า {err} — สูตรอ้างอิงเซลล์ที่ถูกลบหรือหารด้วยศูนย์"))
        check_titles(sh, rep)
        if sh is data_sheet:
            continue
        if not any(isinstance(v, (int, float)) and not isinstance(v, bool) for row in sh.grid for v in row):
            continue
        if check_hourly(sh, rep):
            continue
        if "สรุป" in sh.name and rep.numbers:
            check_summary(sh, rep)
            check_summary_tables(sh, rep)

    # ตรวจเฉพาะช่องที่ใส่สีไว้ในไฟล์
    fills = {s.name: s.fills for s in visible}
    rep.issues = [i for i in rep.issues
                  if i.exempt_color or not i.cell or fills.get(i.sheet, {}).get(i.cell)]
    return rep


# ---------------------------------------------------------------- images / OCR
_XLSX_CACHE: dict = {}


def as_xlsx(data: bytes, filename: str):
    """คืนค่าไฟล์ในรูปแบบ .xlsx (แปลงด้วย LibreOffice ถ้าต้นทางเป็น .xls) — แปลงครั้งเดียวแล้วจำไว้"""
    if not filename.lower().endswith(".xls"):
        return data
    key = hash(data)
    if key in _XLSX_CACHE:
        return _XLSX_CACHE[key]
    out_bytes = None
    if shutil.which("soffice"):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / "in.xls"
            src.write_bytes(data)
            subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", td, str(src)],
                           capture_output=True, timeout=300)
            out = Path(td) / "in.xlsx"
            if out.exists():
                out_bytes = out.read_bytes()
    _XLSX_CACHE[key] = out_bytes
    return out_bytes


def extract_images(data: bytes, filename: str) -> list:
    """ดึงรูปในไฟล์ออกมา (คืนค่า list ของ bytes)"""
    blob = as_xlsx(data, filename)
    if not blob:
        return []
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except Exception:
        return []
    return [z.read(n) for n in sorted(z.namelist()) if n.startswith("xl/media/")]


# ---------------------------------------------------------------- charts
def _range_cells(ref: str) -> list:
    """'กราฟ!$B$43:$B$51,กราฟ!$C$53:$C$60' -> (ชื่อชีต, [(row, col), ...])"""
    sheet, cells = "", []
    for part in ref.split(","):
        part = part.strip()
        if "!" in part:
            sh, rng = part.split("!", 1)
            sheet = sh.strip("'\"")
        else:
            rng = part
        rng = rng.replace("$", "")
        m = re.match(r"^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", rng)
        if not m:
            continue
        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3) or m.group(1), int(m.group(4) or m.group(2))
        col = lambda s: sum((ord(ch) - 64) * 26 ** i for i, ch in enumerate(reversed(s))) - 1
        for r in range(r1, r2 + 1):
            for c in range(col(c1), col(c2) + 1):
                cells.append((r - 1, c))
    return sheet, cells


def read_charts(data: bytes, filename: str) -> list:
    """อ่านกราฟทุกอันในไฟล์: อยู่ชีตไหน อ้างอิงข้อมูลจากชีต/เซลล์ใด และค่าที่กราฟจำไว้เป็นเท่าไร"""
    blob = as_xlsx(data, filename)
    if not blob:
        return []
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except Exception:
        return []
    names = z.namelist()
    wbx = z.read("xl/workbook.xml").decode("utf8", "ignore")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf8", "ignore")
    relmap = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    charts = []
    for nm, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wbx):
        target = "xl/" + relmap.get(rid, "").lstrip("/")
        rp = target.replace("worksheets/", "worksheets/_rels/") + ".rels"
        if rp not in names:
            continue
        for d in re.findall(r'Target="([^"]*drawing\d+\.xml)"', z.read(rp).decode("utf8", "ignore")):
            dp = "xl/" + d.replace("../", "")
            drp = dp.replace("drawings/", "drawings/_rels/") + ".rels"
            if drp not in names:
                continue
            for cpath in re.findall(r'Target="([^"]*chart\d+\.xml)"', z.read(drp).decode("utf8", "ignore")):
                cp = "xl/" + cpath.replace("../", "")
                if cp not in names:
                    continue
                x = z.read(cp).decode("utf8", "ignore")
                title = " ".join(re.findall(r"<a:t>([^<]*)</a:t>", x)[:3])
                for ser in re.findall(r"<c:ser>(.*?)</c:ser>", x, re.S):
                    val = re.search(r"<c:val>(.*?)</c:val>", ser, re.S)
                    if not val:
                        continue
                    ref = re.search(r"<c:f>(.*?)</c:f>", val.group(1), re.S)
                    ref = (ref.group(1) if ref else "").replace("&quot;", '"')
                    src_sheet, cells = _range_cells(ref)
                    pts = [(int(i), float(v)) for i, v in
                           re.findall(r'<c:pt idx="(\d+)"[^>]*>\s*<c:v>([-\d.eE+]+)</c:v>', val.group(1))]
                    charts.append({"sheet": nm, "file": cp.split("/")[-1], "title": title,
                                   "src_sheet": src_sheet, "cells": cells, "points": pts})
    return charts


def check_charts(data: bytes, filename: str, rep: Report) -> list:
    """ตรวจกราฟทีละจุด: ค่าที่กราฟวาดต้องตรงกับเซลล์ต้นทาง และผลรวมต้องตรงกับยอดรวมในชีต data"""
    grids = {s.name: s.grid for s in rep.sheets}
    out = []
    for ch in read_charts(data, filename):
        grid = grids.get(ch["src_sheet"])
        mismatch, checked = [], 0
        if grid:
            for idx, v in ch["points"]:
                if idx >= len(ch["cells"]):
                    continue
                r, c = ch["cells"][idx]
                live = num(grid[r][c]) if r < len(grid) and c < len(grid[r]) else None
                checked += 1
                if live is None or not near(live, v):
                    mismatch.append({"idx": idx + 1, "cell": a1(r, c), "chart": v, "source": live})
        total = sum(v for _, v in ch["points"])
        match = next((k for k in rep.key_totals if near(k["value"], total)), None)
        ch.update({"mismatch": mismatch, "checked": checked, "total": total, "match": match})
        out.append(ch)

        label = f'กราฟในชีต {ch["sheet"]} ({ch["title"] or ch["file"]})'
        for m in mismatch[:20]:
            rep.issues.append(Issue(ch["sheet"], "", "bad", "จุดในกราฟไม่ตรงกับข้อมูลต้นทาง",
                                    f'{label} จุดที่ {m["idx"]} วาดค่า {fmt(m["chart"])} '
                                    f'แต่ {ch["src_sheet"]}!{m["cell"]} เป็น {fmt(m["source"])}'))
        if not mismatch and checked:
            rep.passed.append((ch["sheet"], f'{label} ตรงกับ {ch["src_sheet"]} ครบ {checked} จุด'))
        if match:
            rep.passed.append((ch["sheet"], f'{label} รวมทุกจุดได้ {fmt(total)} ตรงกับ "{match["label"]}" ในชีต data'))
        elif total >= 2:
            rep.issues.append(Issue(ch["sheet"], "", "bad", "ผลรวมของกราฟไม่ตรงกับชีต data",
                                    f'{label} รวมทุกจุดได้ {fmt(total)} แต่ไม่ตรงกับยอดรวมใดในชีต data '
                                    f'({", ".join(fmt(k["value"]) for k in rep.key_totals)})'))
    return out


def read_shapes(data: bytes, filename: str) -> list:
    """อ่านกล่องข้อความ (text box / ลูกศรพร้อมข้อความ) ที่วางทับอยู่บนชีตต่าง ๆ เช่นชีตแผนที่"""
    blob = as_xlsx(data, filename)
    if not blob:
        return []
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except Exception:
        return []
    names = z.namelist()
    wbx = z.read("xl/workbook.xml").decode("utf8", "ignore")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf8", "ignore")
    relmap = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    shapes = []
    for nm, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wbx):
        target = "xl/" + relmap.get(rid, "").lstrip("/")
        rp = target.replace("worksheets/", "worksheets/_rels/") + ".rels"
        if rp not in names:
            continue
        for d in re.findall(r'Target="([^"]*drawing\d+\.xml)"', z.read(rp).decode("utf8", "ignore")):
            dp = "xl/" + d.replace("../", "")
            if dp not in names:
                continue
            x = z.read(dp).decode("utf8", "ignore")
            for sp in re.findall(r"<xdr:sp[ >].*?</xdr:sp>", x, re.S):
                for para in re.findall(r"<a:p>.*?</a:p>", sp, re.S):   # แยกทีละบรรทัด
                    text = "".join(re.findall(r"<a:t>([^<]*)</a:t>", para))
                    text = re.sub(r"\s+", " ", text).strip()
                    if text:
                        shapes.append({"sheet": nm, "text": text})
    return shapes


def _clean_label(t: str) -> str:
    t = re.sub(r"[\d,\.]+", " ", t)
    t = re.sub(r"จำนวน|ยอดรวม|\(|\)|คัน|คน|ทั้งวัน|เฉลี่ย|:|\s", "", t)
    return t


def check_shapes(data: bytes, filename: str, rep: Report) -> list:
    """ตรวจตัวเลขในกล่องข้อความบนชีตแผนที่/สรุป ว่าตรงกับยอดรวมในชีต data หรือไม่"""
    import difflib

    keys = rep.key_totals
    if not keys:
        return []
    # ถ้ารายงานมีหลายวัน ตัวเลขบนแผนที่จะเป็นค่าเฉลี่ย จึงเทียบกับกลุ่ม "เฉลี่ย" ก่อน
    avg = [k for k in keys if "เฉลี่ย" in k["group"]]
    pool = avg or keys
    grand_car = next((k for k in pool if "ทั้งหมด" in k["label"]), None)
    people = [k for k in pool if "คน" in k["label"]]
    grand_people = max(people, key=lambda k: k["value"]) if people else None

    results = []
    for sh in read_shapes(data, filename):
        text = sh["text"]
        body = re.sub(r"\d[\d,]*\.?\d*\s*%", " ", text)     # ตัดค่าเปอร์เซ็นต์ออก
        nums = [float(n.replace(",", "")) for n in re.findall(r"\d[\d,]*\.?\d*", body)]
        nums = [n for n in nums if n >= 2]
        if not nums:
            continue
        label = _clean_label(text)
        best, score = None, 0.0
        if len(label) >= 4:
            for k in pool:
                r = difflib.SequenceMatcher(None, label, _clean_label(k["label"])).ratio()
                if r > score:
                    best, score = k, r
            if score < 0.6:
                best = None
        if best is None:
            if re.match(r"^(รถผ่าน|รถวิ่งผ่าน|รถ)$", label) and grand_car:
                best = grand_car
            elif re.match(r"^(คนผ่าน|คนเดินผ่าน|คน)$", label) and grand_people:
                best = grand_people
        got = min(nums, key=lambda n: abs(n - best["value"])) if best else max(nums)
        if best is None:
            near_by = min(pool, key=lambda k: abs(k["value"] - got))
            if near(got, near_by["value"]) or abs(near_by["value"] - got) <= max(abs(near_by["value"]) * 0.10, 1):
                best = near_by
            else:
                if got < 10:            # ตัวเลขเล็ก ๆ ในข้อความทั่วไป ไม่ใช่ยอดรวม
                    continue
                results.append({**sh, "value": got, "expect": None,
                                "status": "ไม่พบยอดที่ตรงกันในชีต data"})
                rep.issues.append(Issue(
                    sh["sheet"], "", "warn", "กล่องข้อความมียอดที่ไม่มีในชีต data",
                    f'ชีต {sh["sheet"]} · กล่องข้อความ "{text[:60]}" มีตัวเลข {fmt(got)} '
                    f'ซึ่งไม่ตรงกับยอดรวมใดในชีต data — ตรวจว่าเป็นข้อความค้างจากงานเดิมหรือไม่'))
                continue
        if abs(got - best["value"]) > max(abs(best["value"]) * 0.25, 1):
            results.append({**sh, "value": got, "expect": None,
                            "status": "ไม่พบยอดที่ตรงกันในชีต data"})
            rep.issues.append(Issue(
                sh["sheet"], "", "warn", "กล่องข้อความมียอดที่ไม่มีในชีต data",
                f'ชีต {sh["sheet"]} · กล่องข้อความ "{text[:60]}" มีตัวเลข {fmt(got)} '
                f'ซึ่งไม่ตรงกับยอดรวมใดในชีต data — ตรวจว่าเป็นข้อความค้างจากงานเดิมหรือไม่'))
            continue
        ok = near(got, best["value"])
        results.append({**sh, "value": got, "expect": best,
                        "status": "ตรง" if ok else "ไม่ตรง"})
        if ok:
            rep.passed.append((sh["sheet"], f'กล่องข้อความ "{text[:40]}" ตรงกับ {best["label"]} = {fmt(best["value"])}'))
        else:
            rep.issues.append(Issue(
                sh["sheet"], "", "bad", "ตัวเลขในกล่องข้อความไม่ตรงกับชีต data",
                f'ชีต {sh["sheet"]} · กล่องข้อความ "{text[:60]}" ใส่ค่า {fmt(got)} '
                f'แต่ยอดจริงคือ {best["label"]} = {fmt(best["value"])} (ชีต {rep.data_name} เซลล์ {best["cell"]}) '
                f'— ต่าง {fmt(got - best["value"])}'))
    return results


def check_images(data: bytes, filename: str, rep: Report) -> list:
    """ดึงรูปในไฟล์มาแสดงประกอบ (ไม่ตรวจตัวเลขในรูป — ตัวเลขจริงอยู่ในกล่องข้อความ)"""
    return [{"index": i, "image": img} for i, img in enumerate(extract_images(data, filename), 1)]
